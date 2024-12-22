from fastapi import FastAPI, File, UploadFile, HTTPException, Form
from typing import Optional
from pydantic import BaseModel
from dotenv import load_dotenv
from langchain.embeddings import OpenAIEmbeddings
from langchain.text_splitter import CharacterTextSplitter
from langchain_core.prompts import ChatPromptTemplate
from langchain.chains import create_extraction_chain_pydantic


from pymongo import MongoClient
from langchain.chat_models import ChatOpenAI
import openai
import os
import uuid
import numpy as np
import uvicorn
from scipy.linalg import norm
from docx import Document
import fitz  
from PIL import Image
import io
import json
import re
from bson import ObjectId
import openpyxl
from lxml import etree
from bs4 import BeautifulSoup

load_dotenv()

# Configure OpenAI API key
openai.api_key = os.getenv("OPENAI_API_KEY")

# Initialize FastAPI
app = FastAPI()

# MongoDB connection
mongo_client = MongoClient(os.getenv("MONGODB_URI"))
db = mongo_client["onespace"]
collection = db["filedatas"]

# LangChain embedding model setup
embedding_model = OpenAIEmbeddings()

# Text splitter for large files
text_splitter = CharacterTextSplitter(separator="\n", chunk_size=2000, chunk_overlap=200)

@app.get("/")
async def health_check():
    return {"status": "ok", "message": "Service is running"}






def preprocess_text(text):
    """Normalize text without removing special characters."""
    # Convert to lowercase to ensure uniformity
    normalized_text = text.lower()
    # Replace multiple spaces with a single space
    normalized_text = " ".join(normalized_text.split())
    return normalized_text


def extract_text_from_file(file: UploadFile):
    """Extract text based on file type."""
    filename = file.filename.lower()
    mime_type = file.content_type

    # Read file content into memory
    file_content = file.file.read()

    if mime_type == "text/plain" or filename.endswith(".txt"):
        return file_content.decode('utf-8')
    
    elif mime_type == "application/pdf" or filename.endswith(".pdf"):
        doc = fitz.open("pdf", file_content)
        text = ""
        for page in doc:
            text += page.get_text()
        return text

    elif mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document" or filename.endswith(".docx"):
        doc = Document(io.BytesIO(file_content))
        text = "\n".join([para.text for para in doc.paragraphs])
        return text

    elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" or filename.endswith(".xlsx"):
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        text = ""
        # Extract text from each sheet and each row
        for sheet in workbook.sheetnames:
            worksheet = workbook[sheet]
            for row in worksheet.iter_rows(values_only=True):
                text += " ".join([str(cell) if cell is not None else "" for cell in row]) + "\n"
        return text
    # Handle JavaScript (JS) files as plain text
    elif mime_type == "text/javascript" or filename.endswith(".js"):
        return file_content.decode('utf-8')

    # Handle HTML files (use BeautifulSoup for parsing)
    elif mime_type == "text/html" or filename.endswith(".html"):
        soup = BeautifulSoup(file_content, "html.parser")
        return soup.get_text()

    # Handle XML files (use lxml for parsing)
    elif mime_type == "application/xml" or filename.endswith(".xml"):
        try:
            tree = etree.fromstring(file_content)
            return etree.tostring(tree, pretty_print=True, encoding="unicode")
        except etree.XMLSyntaxError as e:
            return f"XML parsing error: {e}"

    # Handle JSON files
    elif mime_type == "application/json" or filename.endswith(".json"):
        try:
            json_data = json.loads(file_content.decode('utf-8'))
            return json.dumps(json_data, indent=4)  # pretty print the JSON
        except json.JSONDecodeError as e:
            raise ValueError(f"JSON parsing error: {e}")

    elif mime_type.startswith("image/") or filename.endswith((".jpg", ".jpeg", ".png")):
        image = Image.open(io.BytesIO(file_content))
        # OCR processing (if necessary) can be added here
        return "Image text processing is not implemented in this example"

    else:
        return f"Unsupported file type: {mime_type}"

class AgenticChunker:
    def __init__(self, openai_api_key=None):
        self.chunks = {}
        self.id_truncate_limit = 5

        # Whether or not to update/refine summaries and titles as you get new information
        self.generate_new_metadata_ind = True
        self.print_logging = True

        if openai_api_key is None:
            openai_api_key = os.getenv("OPENAI_API_KEY")

        if openai_api_key is None:
            raise ValueError("API key is not provided and not found in environment variables")

        self.llm = ChatOpenAI(model='gpt-4-1106-preview', openai_api_key=openai_api_key, temperature=0)
        self.embeddings = OpenAIEmbeddings(openai_api_key=openai_api_key)


    def extract_numeric_details(self, text):
        """
        Extracts all numbers and their contexts to prioritize numerical data
        """
        # Regular expression for capturing numbers (integers, decimals, percentages, etc.)
        number_regex = r'(\d+\.?\d*%?|\.\d+)([^\d%]+)?'
        matches = re.findall(number_regex, text)
        
        # Highlight numeric information in the context of the original text
        numeric_details = " ".join([f"{match[0]} {match[1] if match[1] else ''}" for match in matches])
        return numeric_details.strip()
    
    
    def add_propositions(self, propositions):
        for proposition in propositions:
            self.add_proposition(proposition)
    
    def add_proposition(self, proposition):
        if self.print_logging:
            print (f"\nAdding: '{proposition}'")

        # If it's your first chunk, just make a new chunk and don't check for others
        if len(self.chunks) == 0:
            if self.print_logging:
                print ("No chunks, creating a new one")
            self._create_new_chunk(proposition)
            return

        chunk_id = self._find_relevant_chunk(proposition)

        # If a chunk was found then add the proposition to it
        if chunk_id:
            if self.print_logging:
                print (f"Chunk Found ({self.chunks[chunk_id]['chunk_id']}), adding to: {self.chunks[chunk_id]['title']}")
            self.add_proposition_to_chunk(chunk_id, proposition)
            return
        else:
            if self.print_logging:
                print ("No chunks found")
            # If a chunk wasn't found, then create a new one
            self._create_new_chunk(proposition)
        

    def add_proposition_to_chunk(self, chunk_id, proposition):
        # Add then
        self.chunks[chunk_id]['propositions'].append(proposition)

        # Then grab a new summary
        if self.generate_new_metadata_ind:
            self.chunks[chunk_id]['summary'] = self._update_chunk_summary(self.chunks[chunk_id])
            self.chunks[chunk_id]['title'] = self._update_chunk_title(self.chunks[chunk_id])

    def _update_chunk_summary(self, chunk):
        """
        If you add a new proposition to a chunk, you may want to update the summary or else they could get stale
        """
        PROMPT = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    You are the steward of a group of chunks which represent groups of sentences that talk about a similar topic
                    A new proposition was just added to one of your chunks, you should generate a very brief 1-sentence summary which will inform viewers what a chunk group is about.

                    A good summary will say what the chunk is about, and give any clarifying instructions on what to add to the chunk.

                    You will be given a group of propositions which are in the chunk and the chunks current summary.

                    Your summaries should anticipate generalization. If you get a proposition about apples, generalize it to food.
                    Or month, generalize it to "date and times".
                    

                    Example:
                    Input: Proposition: Greg likes to eat pizza
                    Output: This chunk contains information about the types of food Greg likes to eat.

                    Only respond with the chunk new summary, nothing else.
                    """,
                ),
                ("user", "Chunk's propositions:\n{proposition}\n\nCurrent chunk summary:\n{current_summary}"),
            ]
        )

        runnable = PROMPT | self.llm

        new_chunk_summary = runnable.invoke({
            "proposition": "\n".join(chunk['propositions']),
            "current_summary" : chunk['summary']
        }).content

        return new_chunk_summary
    
    def _update_chunk_title(self, chunk):
        """
        If you add a new proposition to a chunk, you may want to update the title or else it can get stale
        """
        PROMPT = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    You are the steward of a group of chunks which represent groups of sentences that talk about a similar topic
                    A new proposition was just added to one of your chunks, you should generate a very brief updated chunk title which will inform viewers what a chunk group is about.

                    A good title will say what the chunk is about.

                    You will be given a group of propositions which are in the chunk, chunk summary and the chunk title.

                    Your title should anticipate generalization. If you get a proposition about apples, generalize it to food.
                    Or month, generalize it to "date and times".

                    Example:
                    Input: Summary: This chunk is about dates and times that the author talks about
                    Output: Date & Times

                    Only respond with the new chunk title, nothing else.
                    """,
                ),
                ("user", "Chunk's propositions:\n{proposition}\n\nChunk summary:\n{current_summary}\n\nCurrent chunk title:\n{current_title}"),
            ]
        )

        runnable = PROMPT | self.llm

        updated_chunk_title = runnable.invoke({
            "proposition": "\n".join(chunk['propositions']),
            "current_summary" : chunk['summary'],
            "current_title" : chunk['title']
        }).content

        return updated_chunk_title

    def generate_chunk_embeddings(self):
        for chunk_id, chunk in self.chunks.items():
            text = " ".join(chunk['propositions'])
            # Using embed_documents to generate embeddings for each chunk
            chunk_embedding = self.embeddings.embed_documents([text])[0]
            self.chunks[chunk_id]['embedding'] = chunk_embedding
            
    def _get_new_chunk_summary(self, proposition):
        PROMPT = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    You are the steward of a group of chunks which represent groups of sentences that talk about a similar topic
                    You should generate a very brief 1-sentence summary which will inform viewers what a chunk group is about.

                    A good summary will say what the chunk is about, and give any clarifying instructions on what to add to the chunk.

                    You will be given a proposition which will go into a new chunk. This new chunk needs a summary.

                    Your summaries should anticipate generalization. If you get a proposition about apples, generalize it to food.
                    Or month, generalize it to "date and times".

                    Example:
                    Input: Proposition: Greg likes to eat pizza
                    Output: This chunk contains information about the types of food Greg likes to eat.

                    Only respond with the new chunk summary, nothing else.
                    """,
                ),
                ("user", "Determine the summary of the new chunk that this proposition will go into:\n{proposition}"),
            ]
        )

        runnable = PROMPT | self.llm

        new_chunk_summary = runnable.invoke({
            "proposition": proposition
        }).content

        return new_chunk_summary
    
    def _get_new_chunk_title(self, summary):
        PROMPT = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    You are the steward of a group of chunks which represent groups of sentences that talk about a similar topic
                    You should generate a very brief few word chunk title which will inform viewers what a chunk group is about.

                    A good chunk title is brief but encompasses what the chunk is about

                    You will be given a summary of a chunk which needs a title

                    Your titles should anticipate generalization. If you get a proposition about apples, generalize it to food.
                    Or month, generalize it to "date and times".

                    Example:
                    Input: Summary: This chunk is about dates and times that the author talks about
                    Output: Date & Times

                    Only respond with the new chunk title, nothing else.
                    """,
                ),
                ("user", "Determine the title of the chunk that this summary belongs to:\n{summary}"),
            ]
        )

        runnable = PROMPT | self.llm

        new_chunk_title = runnable.invoke({
            "summary": summary
        }).content

        return new_chunk_title


    def _create_new_chunk(self, proposition):
        new_chunk_id = str(uuid.uuid4())[:self.id_truncate_limit] # I don't want long ids
        new_chunk_summary = self._get_new_chunk_summary(proposition)
        new_chunk_title = self._get_new_chunk_title(new_chunk_summary)
        
        
        numeric_details = self.extract_numeric_details(proposition)
    
        # Use the numeric details if any, otherwise use the full proposition
        # We store both in the chunk, not replacing one with the other.
        final_proposition = proposition  # Start with the original proposition
        
        # If numeric details are found, include them
        if numeric_details:
            final_proposition = proposition + " " + numeric_details

        self.chunks[new_chunk_id] = {
            'chunk_id' : new_chunk_id,
            'propositions': [proposition, numeric_details] if numeric_details else [proposition],
            'title' : new_chunk_title,
            'summary': new_chunk_summary,
            'chunk_index' : len(self.chunks)
        }
        if self.print_logging:
            print (f"Created new chunk ({new_chunk_id}): {new_chunk_title}")
    
    def get_chunk_outline(self):
        """
        Get a string which represents the chunks you currently have.
        This will be empty when you first start off
        """
        chunk_outline = ""

        for chunk_id, chunk in self.chunks.items():
            single_chunk_string = f"""Chunk ID: {chunk['chunk_id']}\nChunk Name: {chunk['title']}\nChunk Summary: {chunk['summary']}\n\n"""
        
            chunk_outline += single_chunk_string
        
        return chunk_outline

    def _find_relevant_chunk(self, proposition):
        current_chunk_outline = self.get_chunk_outline()

        PROMPT = ChatPromptTemplate.from_messages(
            [
                (
                    "system",
                    """
                    Determine whether or not the "Proposition" should belong to any of the existing chunks.

                    A proposition should belong to a chunk of their meaning, direction, or intention are similar.
                    The goal is to group similar propositions and chunks.

                    If you think a proposition should be joined with a chunk, return the chunk id.
                    If you do not think an item should be joined with an existing chunk, just return "No chunks"

                    Example:
                    Input:
                        - Proposition: "Greg really likes hamburgers"
                        - Current Chunks:
                            - Chunk ID: 2n4l3d
                            - Chunk Name: Places in San Francisco
                            - Chunk Summary: Overview of the things to do with San Francisco Places

                            - Chunk ID: 93833k
                            - Chunk Name: Food Greg likes
                            - Chunk Summary: Lists of the food and dishes that Greg likes
                    Output: 93833k
                    """,
                ),
                ("user", "Current Chunks:\n--Start of current chunks--\n{current_chunk_outline}\n--End of current chunks--"),
                ("user", "Determine if the following statement should belong to one of the chunks outlined:\n{proposition}"),
            ]
        )

        runnable = PROMPT | self.llm

        chunk_found = runnable.invoke({
            "proposition": proposition,
            "current_chunk_outline": current_chunk_outline
        }).content

        # Pydantic data class
        class ChunkID(BaseModel):
            """Extracting the chunk id"""
            chunk_id: Optional[str]
            
        # Extraction to catch-all LLM responses. This is a bandaid
        extraction_chain = create_extraction_chain_pydantic(pydantic_schema=ChunkID, llm=self.llm)
        extraction_found = extraction_chain.run(chunk_found)
        if extraction_found:
            chunk_found = extraction_found[0].chunk_id

        # If you got a response that isn't the chunk id limit, chances are it's a bad response or it found nothing
        # So return nothing
        if len(chunk_found) != self.id_truncate_limit:
            return None

        return chunk_found
    
    def get_chunks(self, get_type='dict'):
        """
        This function returns the chunks in the format specified by the 'get_type' parameter.
        If 'get_type' is 'dict', it returns the chunks as a dictionary.
        If 'get_type' is 'list_of_strings', it returns the chunks as a list of strings, where each string is a proposition in the chunk.
        """
        if get_type == 'dict':
            return self.chunks
        if get_type == 'list_of_strings':
            chunks = []
            for chunk_id, chunk in self.chunks.items():
                chunks.append(" ".join([x for x in chunk['propositions']]))
            return chunks
    
    def pretty_print_chunks(self):
        print (f"\nYou have {len(self.chunks)} chunks\n")
        for chunk_id, chunk in self.chunks.items():
            print(f"Chunk #{chunk['chunk_index']}")
            print(f"Chunk ID: {chunk_id}")
            print(f"Summary: {chunk['summary']}")
            print(f"Propositions:")
            for prop in chunk['propositions']:
                print(f"    -{prop}")
            print("\n\n")

    def pretty_print_chunk_outline(self):
        print ("Chunk Outline\n")
        print(self.get_chunk_outline())

@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    metadata: Optional[str] = Form(None)
):
    try:
        metadata_dict = json.loads(metadata)
        document_id = metadata_dict["id"]
        # Initialize the AgenticChunker
        agentic_chunker = AgenticChunker(openai_api_key=os.getenv("OPENAI_API_KEY"))
        
        # Extract text from the file
        raw_text = extract_text_from_file(file)
        file_metadata = json.loads(metadata) if metadata else {}
        
        # Preprocess the extracted text (e.g., remove unwanted characters, format)
        processed_text = preprocess_text(raw_text)
        
        filename = file_metadata.get("name", file.filename)
        combined_text = f"Filename: {filename}\n{processed_text}"
        
        # Split the processed text into propositions
        propositions = text_splitter.split_text(combined_text)
        
        # Add each proposition to the AgenticChunker
        agentic_chunker.add_propositions(propositions)
        agentic_chunker.generate_chunk_embeddings()
        agentic_chunker.pretty_print_chunks()
        
        # Store chunks into MongoDB
        new_chunks = []
        
        for chunk_id, chunk in agentic_chunker.get_chunks().items():
            new_chunks.append({
                "filename": file_metadata.get("name", file.filename),
                "chunk_id": chunk_id,
                "title": chunk["title"],
                "summary": chunk["summary"],
                "propositions": chunk["propositions"],
                "embedding": chunk["embedding"],
            })

        # Update existing document in MongoDB
        result = collection.update_one(
            {"doc_id": document_id},  # Filter by document_id
            {
                "$set": {
                    "chunks": new_chunks,
                    "filename": file_metadata.get("name", file.filename),  # Update filename
                    "is_embedded": True
                }
            },
            upsert=True  # Insert document if it doesn't exist
        )
        
        if(result):
            # Return a success message along with chunk overview
            return {
                "message": "File processed and added to MongoDB",
                "success": True,
                "chunk_outline": agentic_chunker.get_chunk_outline(),
            }
        

    except ValueError as ve:
        # Raise an HTTPException for any value-related errors
        raise HTTPException(status_code=400, detail=str(ve))
    
    except Exception as e:
        print(f"error: {e}")
        # Catch any other unexpected errors and raise an HTTPException
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/rank")
async def rank_query(query: str, user_id: str, organization: str):
    try:
        processed_text = preprocess_text(query)
        # Generate embedding for the query using OpenAI embeddings
        query_embedding = openai.Embedding.create(input=processed_text, model="text-embedding-ada-002")['data'][0]['embedding']
        
        # Perform similarity search for chunks in MongoDB by cosine similarity
        query_filter = {"user_id": ObjectId(user_id),"organization": ObjectId(organization),"data.trashed": True}
        documents = collection.find(query_filter)
        
        # Check if documents exist
        if collection.count_documents(query_filter) > 0:
            results = []
            # Iterate over documents in MongoDB
            for doc in documents:
                filename = doc.get("filename")
                chunks = doc.get("chunks", [])
                
                for chunk in chunks:
                    # Skip invalid chunks
                    if not isinstance(chunk.get("embedding"), list):
                        continue

                    # Calculate similarity score
                    embedding = np.array(chunk["embedding"])
                    similarity_score = np.dot(query_embedding, embedding) / (norm(query_embedding) * norm(embedding))
                    
                    # Append result
                    results.append({
                        "filename": filename,
                        "chunk_id": chunk.get("chunk_id"),
                        "title": chunk.get("title"),
                        "summary": chunk.get("summary"),
                        "propositions": chunk.get("propositions"),
                        "data": doc.get("data"),
                        "similarity_score": similarity_score,
                    })

            # Sort and return results
            results = sorted(results, key=lambda x: x["similarity_score"], reverse=True)[:5]
            if not results:
                raise HTTPException(status_code=404, detail="No relevant chunks found.")
            filename = doc.get("filename")
            chunks = doc.get("chunks", [])
            
            # Iterate over each chunk in the document
            for chunk in chunks:
                # Extract chunk embedding (assuming chunk embedding is stored in MongoDB)
                embedding = np.array(chunk["embedding"])
                
                # Calculate cosine similarity between query and chunk embedding
                similarity_score = np.dot(query_embedding, embedding) / (norm(query_embedding) * norm(embedding))
                
                # Append chunk data with similarity score to results
                results.append({
                    "filename": filename,
                    "chunk_id": chunk.get("chunk_id"),
                    "title": chunk.get("title"),
                    "summary": chunk.get("summary"),
                    "propositions": chunk.get("propositions"),
                    "data": doc.get("data"),
                    "similarity_score": similarity_score,
                })
        
            # Sort results by similarity score in descending order and return top results
            results = sorted(results, key=lambda x: x["similarity_score"], reverse=True)[:5]
            
            # Debugging: Log similarity scores to check what’s happening
            print(f"Top similarity scores: {[result['similarity_score'] for result in results]}")
            
            # Generate a prompt for ChatGPT
            prompt = f"""
            **Answer the following question based on the most relevant document. If a document does not directly relate to the query, ignore it:**
    
            ### Question: {query}
            """
            
            # Add chunk content to the prompt
            for i, result in enumerate(results[:3], start=1):
                prompt += f"""
            ### Document {i}:
            **Filename:** {result['filename']}
            **Title:** {result['title']}
            **Summary:** {result['summary']}
            **Propositions:** {result['propositions']}
            """
                
            prompt += """
            ### Provide a concise and direct answer based on the above documents. Provide a concise answer to the question based on the information above in Markdown format.
            """
            print(prompt, "prompt")
            # Get response from ChatGPT,""
            try:
                chat_response = openai.ChatCompletion.create(
                    model="gpt-3.5-turbo",
                    messages=[
                        {"role": "system", "content": "You are a helpful assistant."},
                        {"role": "user", "content": prompt}
                    ]
                )
                answer = chat_response.choices[0].message["content"]
                
                return {"query": query, "answer": answer, "results": results}
        
            except openai.error.OpenAIError as e:
                print(f"Error: {e}")
          
    
            # Retrieve and return the answer from ChatGPT
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
